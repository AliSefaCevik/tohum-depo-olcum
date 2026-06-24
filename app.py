import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os

st.set_page_config(page_title="Tohum Depo Ölçüm Merkezi", page_icon="🌾", layout="centered")

# --- TEK EXCELDE ÇOKLU SAYFA (DEPO) ÜRETEN MOTOR ---
def merkezi_excel_olustur():
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    for d_id in range(1, 13):
        ws = wb.create_sheet(title=f"Depo {d_id}")
        ws.views.sheetView[0].showGridLines = True

        HEADER_FILL = PatternFill(start_color="2B4C7E", end_color="2B4C7E", fill_type="solid")
        SUBHEADER_FILL = PatternFill(start_color="EAF0F6", end_color="EAF0F6", fill_type="solid")
        Z_CELL_FILL = PatternFill(start_color="F4F7FA", end_color="F4F7FA", fill_type="solid")
        ZEBRA_FILL = PatternFill(start_color="F9FBFC", end_color="F9FBFC", fill_type="solid")

        font_title = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        font_sub = Font(name="Arial", size=10, bold=True, color="1A335E")
        font_bold = Font(name="Arial", size=10, bold=True)
        font_regular = Font(name="Arial", size=10)
        font_italic = Font(name="Arial", size=9, italic=True, color="555555")

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center")
        thin_line = Side(border_style="thin", color="D0D7DE")
        thick_line = Side(border_style="medium", color="2B4C7E")
        border_all_thin = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

        ws.merge_cells("A1:G2")
        ws["A1"] = f"DEPO {d_id} - SICAKLIK VE NEM TAKİP RAPORU"
        ws["A1"].font = font_title
        ws["A1"].fill = HEADER_FILL
        ws["A1"].alignment = align_center
        for r in [1, 2]:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = HEADER_FILL

        ws["A3"] = "DEPO / BÖLÜM:"
        ws["B3"] = f"Depo {d_id}"
        ws["C3"] = "ORTAM SICAKLIĞI:"
        ws["E3"] = "ÖLÇÜM SAATİ:"
        ws["G3"] = "NEM (%):"
        for col in ["A", "C", "E", "G"]:
            ws[f"{col}3"].font = font_sub
            ws[f"{col}3"].alignment = align_right
        ws["B3"].font = font_bold
        ws["B3"].alignment = align_center

        ws.merge_cells("A4:G4")
        ws["A4"] = "YIĞIN DİP SICAKLIKLARI GÖRSEL DAĞILIM ŞEMASI (Z)"
        ws["A4"].font = font_sub
        ws["A4"].fill = SUBHEADER_FILL
        ws["A4"].alignment = align_center
        for c in range(1, 8):
            ws.cell(row=4, column=c).fill = SUBHEADER_FILL

        z_points = {
            "B6": "1. ÜST SOL", "D6": "2. ÜST ORTA", "F6": "3. ÜST SAĞ",
            "D9": "4. TAM ORTA",
            "B12": "5. ALT SOL", "D12": "6. ALT ORTA", "F12": "7. ALT SAĞ"
        }
        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}5"].border = Border(top=thick_line)
            ws[f"{col}12"].border = Border(bottom=thick_line)

        for cell_ref, label in z_points.items():
            ws[cell_ref] = 0.0
            ws[cell_ref].font = font_bold
            ws[cell_ref].alignment = align_center
            ws[cell_ref].fill = Z_CELL_FILL
            ws[cell_ref].border = border_all_thin
            
            l_cell = ws[f"{cell_ref[0]}{int(cell_ref[1:])-1}"]
            l_cell.value = label
            l_cell.font = font_italic
            l_cell.alignment = align_center

        ws.merge_cells("A14:G14")
        ws["A14"] = "ÖLÇÜM VERİLERİ ÖZET LİSTESİ"
        ws["A14"].font = font_sub
        ws["A14"].fill = SUBHEADER_FILL
        ws["A14"].alignment = align_center
        for c in range(1, 8):
            ws.cell(row=14, column=c).fill = SUBHEADER_FILL

        headers = ["Nokta No / Tanımı", "Ölçülen Değer (°C)", "Durum / Limit", "Tarih", "Saat", "Ölçen Personel", "Açıklama"]
        for idx, h in enumerate(headers, start=1):
            c = ws.cell(row=15, column=idx)
            c.value = h
            c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.fill = HEADER_FILL
            c.alignment = align_center

        point_names = [
            "Ortam Sıcaklığı", "1. Üst Sol Noktası", "2. Üst Orta Noktası", "3. Üst Sağ Noktası",
            "4. Tam Orta Noktası", "5. Alt Sol Noktası", "6. Alt Orta Noktası", "7. Alt Sağ Noktası"
        ]
        for idx, name in enumerate(point_names, start=16):
            ws.cell(row=idx, column=1, value=name).font = font_bold
            ws.cell(row=idx, column=2, value="-")
            ws.cell(row=idx, column=3, value="Normal")
            ws.cell(row=idx, column=4, value="-")
            ws.cell(row=idx, column=5, value="-")
            ws.cell(row=idx, column=6, value="-")
            ws.cell(row=idx, column=7, value="Z-Şeması Takibi")
            
            row_fill = ZEBRA_FILL if idx % 2 == 0 else PatternFill(fill_type=None)
            for col in range(1, 8):
                cell = ws.cell(row=idx, column=col)
                cell.border = border_all_thin
                if row_fill.fill_type:
                    cell.fill = row_fill
                cell.font = font_regular

        widths = {"A": 22, "B": 15, "C": 18, "D": 15, "E": 15, "F": 15, "G": 16}
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

    wb.save("merkezi_sablon.xlsx")

if not os.path.exists("merkezi_sablon.xlsx"):
    merkezi_excel_olustur()

# --- STREAMLIT SESSİON STATE YAPISI ---
if 'fabrika_verisi' not in st.session_state:
    st.session_state.fabrika_verisi = {f"Depo {i}": {} for i in range(1, 13)}

st.title("🌾 Merkezi Tohum Ölçüm İstasyonu")
st.write("Tüm depolar tek bir Excel dosyasında toplanıyor kanka!")

# --- 1. ADIM: GENEL BİLGİLER ---
st.subheader("1. Genel Bilgiler")
col_a, col_b = st.columns(2)
with col_a:
    # İŞTE BURADAKİ PARANTEZ HATASINI DÜZELTTİM KANKA, ARTIK LİSTE İÇERİDE:
    depo_no = st.selectbox("Hangi Depoyu Ölçüyorsunuz?", [f"Depo {i}" for i in range(1, 13)])
    tohum_cinsi = st.selectbox("Tohum Cinsi", ["Buğday", "Arpa", "Mısır", "Ayçiçeği", "Nohut"])
with col_b:
    olcen_kisi = st.text_input("Ölçen Kişi", "Ali Sefa")
    saat_araligi = st.text_input("Saat Aralığı", datetime.now().strftime("%H:%M"))

st.markdown("---")

# --- 2. ADIM: VERİ GİRİŞ ALANI ---
st.subheader(f"2. {depo_no} Sıcaklık ve Nem Değerleri")

noktalar = [
    "Ortam Sıcaklığı", "Ortam Nemi (%)",
    "1. Üst Sol (Z)", "2. Üst Orta (Z)", "3. Üst Sağ (Z)", 
    "4. TAM ORTA (Z)", 
    "5. Alt Sol (Z)", "6. Alt Orta (Z)", "7. Alt Sağ (Z)"
]

secilen_nokta = st.selectbox("Nokta Seçin:", noktalar)

if "Nemi" in secilen_nokta:
    deger = st.number_input(f"{secilen_nokta} Değeri", min_value=0, max_value=100, value=35, step=1)
else:
    deger = st.number_input(f"{secilen_nokta} Değeri (°C)", min_value=15.0, max_value=45.0, value=25.0, step=0.1, format="%.1f")

if st.button(f"📌 {depo_no} Hafızasına Kaydet", use_container_width=True):
    st.session_state.fabrika_verisi[depo_no][secilen_nokta] = deger
    st.success(f"✔️ {depo_no} - {secilen_nokta} için {deger} hafızaya alındı!")

# Mevcut Seçili Deponun Durumu
st.markdown(f"### 📊 {depo_no} Giriş Durumu")
for n in noktalar:
    if n in st.session_state.fabrika_verisi[depo_no]:
        birim = " %" if "Nemi" in n else " °C"
        st.write(f"✅ **{n}:** {st.session_state.fabrika_verisi[depo_no][n]}{birim}")
    else:
        st.write(f"❌ **{n}:** Veri bekleniyor...")

st.markdown("---")

# --- 3. ADIM: TÜM DEPOLARI EXCEL'E AKTAR VE İNDİR ---
st.subheader("3. Fabrika Raporunu Kapat")
st.write("Verileri girdiğiniz tüm depolar tek bir dosyada birleşir.")

if st.button("💾 TÜM DEPOLARI TEK EXCEL'E AKTAR VE RAPORU İNDİR", use_container_width=True):
    wb = openpyxl.load_workbook("merkezi_sablon.xlsx")
    tarih_str = datetime.now().strftime("%d.%m.%Y")
    
    kirmizi_dolgu = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    kirmizi_yazi = Font(name="Arial", size=10, bold=True, color="9C0006")
    
    aktif_depo_sayisi = 0
    
    for d_name, d_data in st.session_state.fabrika_verisi.items():
        if len(d_data) == len(noktalar):
            aktif_depo_sayisi += 1
            ws = wb[d_name]
            
            ws["D3"] = float(d_data["Ortam Sıcaklığı"])
            ws["F3"] = saat_araligi
            ws["G3"] = f"NEM: %{d_data['Ortam Nemi (%)']}"
            
            ws["B6"] = float(d_data["1. Üst Sol (Z)"])
            ws["D6"] = float(d_data["2. Üst Orta (Z)"])
            ws["F6"] = float(d_data["3. Üst Sağ (Z)"])
            ws["D9"] = float(d_data["4. TAM ORTA (Z)"])
            ws["B12"] = float(d_data["5. Alt Sol (Z)"])
            ws["D12"] = float(d_data["6. Alt Orta (Z)"])
            ws["F12"] = float(d_data["7. Alt Sağ (Z)"])
            
            dip_noktalari = [
                ("1. Üst Sol Noktası", float(d_data["1. Üst Sol (Z)"]), 17),
                ("2. Üst Orta Noktası", float(d_data["2. Üst Orta (Z)"]), 18),
                ("3. Üst Sağ Noktası", float(d_data["3. Üst Sağ (Z)"]), 19),
                ("4. Tam Orta Noktası", float(d_data["4. TAM ORTA (Z)"]), 20),
                ("5. Alt Sol Noktası", float(d_data["5. Alt Sol (Z)"]), 21),
                ("6. Alt Orta Noktası", float(d_data["6. Alt Orta (Z)"]), 22),
                ("7. Alt Sağ Noktası", float(d_data["7. Alt Sağ (Z)"]), 23)
            ]
            
            en_yuksek_deger = max([x[1] for x in dip_noktalari])
            
            mapping = {
                16: ("Ortam Sıcaklığı", "Ortam Sıcaklığı"),
                17: ("1. Üst Sol Noktası", "1. Üst Sol (Z)"),
                18: ("2. Üst Orta Noktası", "2. Üst Orta (Z)"),
                19: ("3. Üst Sağ Noktası", "3. Üst Sağ (Z)"),
                20: ("4. Tam Orta Noktası", "4. TAM ORTA (Z)"),
                21: ("5. Alt Sol Noktası", "5. Alt Sol (Z)"),
                22: ("6. Alt Orta Noktası", "6. Alt Orta (Z)"),
                23: ("7. Alt Sağ Noktası", "7. Alt Sağ (Z)")
            }
            
            for row_idx, (label, state_key) in mapping.items():
                guncel_deger = float(d_data[state_key])
                ws.cell(row=row_idx, column=2, value=guncel_deger)
                ws.cell(row=row_idx, column=4, value=tarih_str)
                ws.cell(row=row_idx, column=5, value=saat_araligi)
                ws.cell(row=row_idx, column=6, value=olcen_kisi)
                
                if row_idx != 16 and guncel_deger == en_yuksek_deger:
                    ws.cell(row=row_idx, column=3, value="Yüksek")
                    ws.cell(row=row_idx, column=3).fill = kirmizi_dolgu
                    ws.cell(row=row_idx, column=3).font = kirmizi_yazi
                    ws.cell(row=row_idx, column=2).fill = kirmizi_dolgu
                    ws.cell(row=row_idx, column=2).font = kirmizi_yazi

    if aktif_depo_sayisi == 0:
        st.error("En az 1 deponun tüm ölçümlerini tam doldurup hafızaya kaydetmelisin kanka!")
    else:
        cikti_adi = f"Fabrika_Merkezi_Olcum_Raporu_{tarih_str.replace('.', '_')}.xlsx"
        wb.save(cikti_adi)
        
        with open(cikti_adi, "rb") as file:
            st.download_button(
                label=f"📥 {aktif_depo_sayisi} DEPOLUK MERKEZİ RAPORU İNDİR",
                data=file,
                file_name=cikti_adi,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        st.balloons()
